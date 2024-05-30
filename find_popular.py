from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook

# 엑셀 파일 로드
workbook = load_workbook(filename="brand.xlsx")

sheet = workbook.active

header = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36',
}

chrome_options = Options()
#브라우저 꺼짐 방지
chrome_options.add_experimental_option("detach", True)
#불필요한 에러 메시지 없애기
chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"]) # 셀레니움 로그 무시

chrome_options.add_argument(r'load-extension=C:\Users\jaewon\AppData\Local\Google\Chrome Dev\User Data\Profile 1\Extensions\cgococegfcmmfcjggpgelfbjkkncclkf\1.1.9.3_0')

service = webdriver.ChromeService(executable_path='C:\Program Files\Google\Chrome Dev\Application\chrome.exe')


find_site = ""


if __name__ == '__main__':
    
    b = webdriver.Chrome(options=chrome_options)
    
    b.get(find_site)
    
    btn = b.find_element(By.XPATH, "//*[contains(text(), 'Brands')]")
    
    b.execute_script("arguments[0].click();", btn)
    
    contains = b.find_elements(By.XPATH, f"//*[contains(@id,'section-')]")
    
    row = 1
    
    print(f"총 길이 = {len(contains)}")
    
    for i, c in enumerate(contains):
        t = c.find_elements(By.XPATH, f"//*[contains(@id,'ern:brand::')]")
        
        print(f"현재 = {i}")
        
        # for a in t:
        for a in t:
            z = a.find_element(By.XPATH, './*')
            sheet.cell(row=row, column=1).value = z.text
            row += 1
            
            
    workbook.save(filename="brand.xlsx")        
    print("======끝==========")
    

