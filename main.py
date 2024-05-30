from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import load_workbook, Workbook
from selenium.webdriver.support.ui import Select

# 엑셀 파일 로드
l_workbook = load_workbook(filename="test.xlsx")

sheet = l_workbook.active

header = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36',
}

chrome_options = Options()

chrome_options.add_experimental_option("detach", True)

# 불필요한 에러 메시지 없애기
chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])

#브라우저 꺼짐 방지
# chrome_options.add_experimental_option("detach", True)
#불필요한 에러 메시지 없애기
# chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"]) # 셀레니움 로그 무시

# chrome_options.add_argument(r'load-extension=C:\Users\jaewon\AppData\Local\Google\Chrome Dev\User Data\Profile 1\Extensions\cgococegfcmmfcjggpgelfbjkkncclkf\1.1.9.3_0')

# service = webdriver.ChromeService(executable_path='C:\Program Files\Google\Chrome Dev\Application\chrome.exe')

service = webdriver.ChromeService(executable_path='C:\chrome-win64\chrome.exe')

# =============================클롤링 변수 =================
data_start_row = 1

url = ""

attribute = ""
attribute_value = ""

option_attrivute = ""
option_container = ""

product_detail_type = ""
# =============================클롤링 변수 =================

def remove_none_img(imgs):
        filer_img = []
        
        for img in imgs:
            if img.get_attribute("src") != None:
                filer_img.append(img)
                
        return filer_img
    
def find_imgs(tag):
    imgs = tag.find_elements(By.TAG_NAME, 'img')
    
    filer_img = []
    
    for img in imgs:
        if img.get_attribute("src") != None:
            filer_img.append(img)
    
    return filer_img

def find_imgs_container(img_elements):
    
    for img_tag in img_elements:
        parent_tag = img_tag.find_element(By.XPATH, '..')
        parent_tag_imgs = find_imgs(parent_tag)
        
        while len(parent_tag_imgs) < len(img_elements):
            parent_tag = parent_tag.find_element(By.XPATH, '..')
            parent_tag_imgs = find_imgs(parent_tag)
        
        return parent_tag.find_elements(By.XPATH, './*')

def filter_button(text):
    
    filter_list = ["상품", "사이즈", "size", "선택", "고시", "장바구니"]
    
    result = any(keyword in text for keyword in filter_list) or text == ""
    
    return result

def celar_text(text):
    return text.replace("\n", "").replace("\t", "").replace("\r", "").strip()

def check_color(element):

    color = element.value_of_css_property("color")
    
    rgb = tuple(map(int, color.strip("rgba()").split(",")))
    
    is_gray = all(rgb[i] >= 101 for i in range(3))
    
    return is_gray

def find_text_elements(container, text):
    return container.find_element(By.XPATH, f".//*[contains(text(), '{text}')]")
    

def movement(title):    
    # grid grid-cols-6
    
    # 렌더링 대기
    time.sleep(2)
    
    # 옵션 컨테이너
    try:
        container =  b.find_element(By.XPATH, f"//*[@{option_attrivute}='{option_container}']")
        idontKnowWhyThisDo = container.find_elements(By.XPATH, "./*")
    except Exception as e:
        print("상품 설정 데이터 이외의 상품")
        b.back()
        return
       
    품절옵션 = []
    
    품절옵션.append(title)
    품절옵션.append(b.current_url)
    
    # select 태그 존재 여부 체크 
    
    # selects_tags = container.find_elements(By.XPATH, "//select")
    selects_tags = []
    
    
    if product_detail_type == "select":
        
        for t in idontKnowWhyThisDo:
            try:
                select = t.find_element(By.XPATH, ".//select")
                selects_tags.append(select)
            except Exception as e:
                print("")
    
        # select 태그가 있을 경우 option 데이터 추출
        if len(selects_tags) > 0:
        
            for select in selects_tags:
                
                b.execute_script("arguments[0].style.display = 'block';", select)
            
            options = container.find_elements(By.XPATH, f".//option")
            
            for option in options:
                text = celar_text(option.text)
                
                if text != "" and "선택" not in text:
                    품절옵션.append(text)
        
        # select 태그 없이 구현되어 있을 경우
        else:
            
            for t in idontKnowWhyThisDo:
                try:
                    b.execute_script("arguments[0].click();", t)
                    
                    time.sleep(1)
                    
                    # temp__container =  b.find_element(By.XPATH, f"//*[@{option_attrivute}='{option_container}']")
                    temp__container = b.find_element(By.XPATH, f"//*[contains(@{option_attrivute},'{option_container}')]")
                    
                    
                    size_list = temp__container.text.strip().split("\n")
                    
                    for size in size_list:
                        size_element = temp__container.find_element(By.XPATH, f".//*[contains(text(), '{size}')]")
                        
                        is_gray = check_color(size_element)
                        
                        if is_gray:
                            품절옵션.append("품절")
                        else:
                            품절옵션.append(size_element.text)
                                          
                except Exception as e:
                    print("")
    
    else:
        print("not select")
         
        # select 태그가 경을 경우 button 데이터 추출
        for child in idontKnowWhyThisDo:

            try:
                option = child.find_element(By.XPATH, './/button')
                
                check_button = filter_button(option.text)
                
                if check_button:
                    continue
                
                is_gray = False
            
                color = option.value_of_css_property("color")
                
                rgb = tuple(map(int, color.strip("rgba()").split(",")))
                
                is_gray = all(rgb[i] >= 101 for i in range(3))
            

                if is_gray:
                    품절옵션.append("품절")
                else:
                    품절옵션.append(option.text)
                    
            except Exception as e:
                print("button 없음")
                품절옵션.append(child.text)
            
        
    data_sheet.append(품절옵션)
    
    print(품절옵션)
    data_excel.save(filename=f"{file_name}.xlsx")
    
    # 페이지 뒤로가기
    b.back()
    

def getContainer():
    return b.find_element(By.XPATH, f"//*[contains(@{attribute},'{attribute_value}')]")
    


def scroll_down():
    # 페이지 높이 구하기
    last_height = b.execute_script("return document.body.scrollHeight")

    b.find_element(By.TAG_NAME, "body").send_keys(Keys.END)
    
    # 스크롤 후 대기
    time.sleep(2)
    
    # 새로운 페이지 높이 구하기
    new_height = b.execute_script("return document.body.scrollHeight")
    
    if new_height == last_height:
        
        # 끝에 도달했을 때 True 리턴
        return True
    
    last_height = new_height
        
def product_search(search_product):
    
    try:
    
        while True:
            
            time.sleep(1)
                
            imgs = find_imgs(getContainer())
            imgs_container = find_imgs_container(imgs)
            
            product = ""
             
            for item in imgs_container:
                if item.text not in search_product:
                    product = item
                    break
                
            if product == "":
                break
            
            
            title = product.text
            search_product.append(product.text)
            
            click_element = product.find_element(By.TAG_NAME, 'img')
            
            b.execute_script("arguments[0].click();", click_element)
            
            print(title)
            movement(title.replace("\n", ""))
            
    except Exception as e:
        
        print("컨테이너 검색 끝")
                
   

data_excel = ""
data_sheet = ""
file_name = ""

if __name__ == '__main__':
    
    
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            
            # 2번쨰 Column이 o이면 실행, x이면 실행 안함
            if row[1] == "x":
                continue
            
            # 엑셀 파일 생성
            data_excel = Workbook()
            data_sheet = data_excel.active
            
            
            if i > data_start_row:
            # if i == 5:
            
                print("Row", i, ":", row)
                
                url = row[0]
                
                if row[3]:
                    attribute = "id"
                    attribute_value = row[3]
                else:
                    attribute = "class"
                    attribute_value = row[4]
                    
                if row[5]:
                    option_attrivute = "id"
                    option_container = row[5]
                else:
                    option_attrivute = "class"
                    option_container = row[6]
                    
                product_detail_type = row[2]
                    
                b = webdriver.Chrome(options=chrome_options)
                b.get(f"{url}")

                data_sheet.append([url])
                file_name = url.replace("https://", "").replace("http://", "").replace(".html", "").replace("/", "").replace("?", "")

            try:
                
                search_product = []
                
                while True:
                    
                    product_search(search_product)
                    res = scroll_down()
                    
                    if res:
                        print(f"========================={url} = 끝=======================================")
                        break
                    
            except Exception as e:
                print(e)
                    
                    
            # WebDriver를 닫습니다.
            b.quit()
