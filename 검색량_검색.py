from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from time import sleep
import re

# ============================================================
# 환경 변수

start_row = 1
엑셀_파일_이름 = "janabrand.xlsx"

크롬_파일_경로 = "C:\Program Files\Google\Chrome Dev\Application\chrome.exe"

# ============================================================

# 엑셀 파일 로드
workbook = load_workbook(filename=f"{엑셀_파일_이름}")

sheet = workbook.active

header = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36',
}

chrome_options = Options()
#브라우저 꺼짐 방지
chrome_options.add_experimental_option("detach", True)
#불필요한 에러 메시지 없애기
chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"]) # 셀레니움 로그 무시

chrome_options.add_argument(r'load-extension=C:\Users\jaewon\AppData\Local\Google\Chrome Dev\User Data\Profile 1\Extensions\cgococegfcmmfcjggpgelfbjkkncclkf\1.2.0.5_0')

service = webdriver.ChromeService(executable_path=f"{크롬_파일_경로}")


if __name__ == '__main__':
    
    b = webdriver.Chrome(options=chrome_options)
    
    max_row_cnt = sheet.max_row
    
    
    for i in range(start_row, max_row_cnt - 1):
        
        keyword = sheet[f'A{i}'].value
        
        keyword = re.sub(r'\(.*?\)', '', keyword)
        
        b.get(f"https://search.shopping.naver.com/search/all?query={keyword.strip()}")
        
        try:
            iframe = b.find_element(By.CSS_SELECTOR, 'iframe#sellerlife-keyword-analysis-frame')
            b.switch_to.frame(iframe)
        except Exception as e:
            print(e)
            continue
        
        # 렌더링 타이머
        sleep(3)
        
        tt = b.find_element(By.CLASS_NAME, "totalprdcnt")
        총상품수 = int(b.find_element(By.CLASS_NAME, "cnt").text.strip().replace(",", "").replace("개", ""))
        
        gg = b.find_element(By.CLASS_NAME, "ovsprdcnt")
        해외상품수 = int(gg.find_element(By.CLASS_NAME, "cnt").text.strip().replace(",", "").replace("개", ""))
        
        판매건수 = b.find_elements(By.XPATH, "//*[contains(@class,'sellcount')]")
        
        top_40_6개월_판매건수 = int(판매건수[0].text.strip().replace(",", "").replace("개", ""))
        top_80_6개월_판매건수 = int(판매건수[1].text.strip().replace(",", "").replace("개", ""))
        
        매출액 = b.find_elements(By.XPATH, "//*[contains(@class,'revenue')]")
        
        top_40_6개월_매출액 = int(매출액[0].text.strip().replace(",", "").replace("원", ""))
        top_80_6개월_매출액 = int(매출액[1].text.strip().replace(",", "").replace("원", ""))
        
        try:
            검색량_최근1달 = int(b.find_element(By.XPATH, "//*[contains(@class,'monthlyQcCnt')]").text.strip().replace(",", ""))
        except Exception as e:
            검색량_최근1달 = b.find_element(By.XPATH, "//*[contains(@class,'monthlyQcCnt')]").text.strip()
        
        try:
            검색량_예상1달 = int(b.find_element(By.XPATH, "//*[contains(@class,'estQcCnt')]").text.strip().replace(",", ""))
        except Exception as e:
            검색량_예상1달 = b.find_element(By.XPATH, "//*[contains(@class,'estQcCnt')]").text

        try:        
            검색량_예상3달평균 = int(b.find_element(By.XPATH, "//*[contains(@class,'estThrQcCnt')]").text.strip().replace(",", ""))
        except Exception as e:
            검색량_예상3달평균 = b.find_element(By.XPATH, "//*[contains(@class,'estQcCnt')]").text
    
        # ===========================클롱링 데이터 =================================
        
        

        sheet.cell(row=i, column=3).value = 총상품수
        sheet.cell(row=i, column=4).value = 해외상품수
        sheet.cell(row=i, column=5).value = 검색량_최근1달
        sheet.cell(row=i, column=6).value = 검색량_예상1달
        sheet.cell(row=i, column=7).value = 검색량_예상3달평균
        sheet.cell(row=i, column=8).value = top_40_6개월_판매건수
        sheet.cell(row=i, column=9).value = top_80_6개월_판매건수
        sheet.cell(row=i, column=10).value = top_40_6개월_매출액
        sheet.cell(row=i, column=11).value = top_80_6개월_매출액
        
        workbook.save(filename=f"{엑셀_파일_이름}")
        print(f"{i} 번째 {sheet[f'A{i}'].value} 브랜드 검색 완료")
        
    print("==================================끝================================================")
    b.quit()
